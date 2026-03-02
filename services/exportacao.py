"""
Serviço de exportação para PDF e Excel
"""
from typing import List, Dict
from datetime import datetime
from decimal import Decimal
from io import BytesIO

# PDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_RIGHT

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class ExportacaoService:
    """Serviço para exportação de relatórios"""
    
    @staticmethod
    def formatar_valor(valor) -> str:
        """Formata valor para exibição"""
        if valor is None:
            return "R$ 0,00"
        if isinstance(valor, (int, float, Decimal)):
            return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        return str(valor)
    
    @staticmethod
    def formatar_data(data) -> str:
        """Formata data para exibição"""
        if data is None:
            return "-"
        if isinstance(data, str):
            return data
        return data.strftime('%d/%m/%Y')
    
    @staticmethod
    def exportar_contas_pagar_pdf(contas: List[Dict], filtros: Dict = None) -> BytesIO:
        """Exporta contas a pagar para PDF"""
        buffer = BytesIO()
        
        # Criar documento PDF em paisagem
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                               rightMargin=1*cm, leftMargin=1*cm,
                               topMargin=1.5*cm, bottomMargin=1*cm)
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#6B46C1'),
            spaceAfter=20,
            alignment=TA_CENTER
        )
        
        elements.append(Paragraph("Relatório de Contas a Pagar", title_style))
        elements.append(Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", 
                                 styles['Normal']))
        elements.append(Spacer(1, 0.5*cm))
        
        # Filtros aplicados
        if filtros:
            filtros_text = "Filtros: "
            if filtros.get('status'):
                filtros_text += f"Status: {filtros['status']} | "
            if filtros.get('data_inicio'):
                filtros_text += f"Período: {filtros['data_inicio']} a {filtros.get('data_fim', 'hoje')} | "
            elements.append(Paragraph(filtros_text, styles['Normal']))
            elements.append(Spacer(1, 0.3*cm))
        
        # Tabela de dados
        data = [['Vencimento', 'Fornecedor', 'Descrição', 'Valor', 'Status']]
        
        total = Decimal('0')
        for conta in contas:
            data.append([
                ExportacaoService.formatar_data(conta.get('data_vencimento')),
                conta.get('fornecedor_fantasia') or conta.get('fornecedor_nome', ''),
                conta.get('descricao', '')[:40],
                ExportacaoService.formatar_valor(conta.get('valor_total')),
                conta.get('status', '').upper()
            ])
            if conta.get('valor_total'):
                total += Decimal(str(conta['valor_total']))
        
        # Linha de total
        data.append(['', '', 'TOTAL:', ExportacaoService.formatar_valor(total), ''])
        
        # Criar tabela
        table = Table(data, colWidths=[3*cm, 6*cm, 8*cm, 3.5*cm, 2.5*cm])
        table.setStyle(TableStyle([
            # Cabeçalho
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6B46C1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Dados
            ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -2), 8),
            ('ALIGN', (3, 1), (3, -1), 'RIGHT'),
            ('ALIGN', (4, 1), (4, -2), 'CENTER'),
            ('GRID', (0, 0), (-1, -2), 0.5, colors.grey),
            
            # Total
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F3F4F6')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 10),
            ('ALIGN', (2, -1), (2, -1), 'RIGHT'),
            ('ALIGN', (3, -1), (3, -1), 'RIGHT'),
        ]))
        
        elements.append(table)
        
        # Construir PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def exportar_contas_pagar_excel(contas: List[Dict], filtros: Dict = None) -> BytesIO:
        """Exporta contas a pagar para Excel"""
        buffer = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Contas a Pagar"
        
        # Estilos
        header_fill = PatternFill(start_color="6B46C1", end_color="6B46C1", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        total_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        total_font = Font(bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws['A1'] = "Relatório de Contas a Pagar"
        ws['A1'].font = Font(bold=True, size=14, color="6B46C1")
        ws.merge_cells('A1:H1')
        
        ws['A2'] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws.merge_cells('A2:H2')
        
        # Cabeçalhos
        headers = ['Data Vencimento', 'Fornecedor', 'Descrição', 'Nº Documento', 
                  'Valor Total', 'Status', 'Filial', 'Centro de Custo']
        row = 4
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Dados
        row = 5
        total = Decimal('0')
        
        for conta in contas:
            ws.cell(row=row, column=1, value=ExportacaoService.formatar_data(conta.get('data_vencimento')))
            ws.cell(row=row, column=2, value=conta.get('fornecedor_fantasia') or conta.get('fornecedor_nome'))
            ws.cell(row=row, column=3, value=conta.get('descricao'))
            ws.cell(row=row, column=4, value=conta.get('numero_documento') or '')
            
            valor = conta.get('valor_total', 0)
            ws.cell(row=row, column=5, value=float(valor) if valor else 0)
            ws.cell(row=row, column=5).number_format = 'R$ #,##0.00'
            
            ws.cell(row=row, column=6, value=conta.get('status', '').upper())
            ws.cell(row=row, column=7, value=conta.get('filial_nome') or '')
            ws.cell(row=row, column=8, value=conta.get('centro_custo_descricao') or '')
            
            # Aplicar bordas
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = border
                ws.cell(row=row, column=col).alignment = Alignment(vertical='center')
            
            if valor:
                total += Decimal(str(valor))
            row += 1
        
        # Total
        ws.cell(row=row, column=4, value='TOTAL:')
        ws.cell(row=row, column=4).font = total_font
        ws.cell(row=row, column=4).fill = total_fill
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='right')
        
        ws.cell(row=row, column=5, value=float(total))
        ws.cell(row=row, column=5).number_format = 'R$ #,##0.00'
        ws.cell(row=row, column=5).font = total_font
        ws.cell(row=row, column=5).fill = total_fill
        
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = border
        
        # Ajustar larguras
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 20
        
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def exportar_contas_receber_pdf(contas: List[Dict], filtros: Dict = None) -> BytesIO:
        """Exporta contas a receber para PDF"""
        buffer = BytesIO()
        
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                               rightMargin=1*cm, leftMargin=1*cm,
                               topMargin=1.5*cm, bottomMargin=1*cm)
        
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#6B46C1'),
            spaceAfter=20,
            alignment=TA_CENTER
        )
        
        elements.append(Paragraph("Relatório de Contas a Receber", title_style))
        elements.append(Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", 
                                 styles['Normal']))
        elements.append(Spacer(1, 0.5*cm))
        
        # Tabela de dados
        data = [['Vencimento', 'Cliente', 'Descrição', 'Valor', 'Status']]
        
        total = Decimal('0')
        for conta in contas:
            data.append([
                ExportacaoService.formatar_data(conta.get('data_vencimento')),
                conta.get('cliente_fantasia') or conta.get('cliente_nome', ''),
                conta.get('descricao', '')[:40],
                ExportacaoService.formatar_valor(conta.get('valor_total')),
                conta.get('status', '').upper()
            ])
            if conta.get('valor_total'):
                total += Decimal(str(conta['valor_total']))
        
        data.append(['', '', 'TOTAL:', ExportacaoService.formatar_valor(total), ''])
        
        table = Table(data, colWidths=[3*cm, 6*cm, 8*cm, 3.5*cm, 2.5*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6B46C1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -2), 8),
            ('ALIGN', (3, 1), (3, -1), 'RIGHT'),
            ('ALIGN', (4, 1), (4, -2), 'CENTER'),
            ('GRID', (0, 0), (-1, -2), 0.5, colors.grey),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F3F4F6')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 10),
            ('ALIGN', (2, -1), (2, -1), 'RIGHT'),
            ('ALIGN', (3, -1), (3, -1), 'RIGHT'),
        ]))
        
        elements.append(table)
        doc.build(elements)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def exportar_contas_receber_excel(contas: List[Dict], filtros: Dict = None) -> BytesIO:
        """Exporta contas a receber para Excel"""
        buffer = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Contas a Receber"
        
        header_fill = PatternFill(start_color="6B46C1", end_color="6B46C1", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        total_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        total_font = Font(bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        ws['A1'] = "Relatório de Contas a Receber"
        ws['A1'].font = Font(bold=True, size=14, color="6B46C1")
        ws.merge_cells('A1:H1')
        
        ws['A2'] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws.merge_cells('A2:H2')
        
        headers = ['Data Vencimento', 'Cliente', 'Descrição', 'Nº Documento', 
                  'Valor Total', 'Status', 'Filial', 'Centro de Custo']
        row = 4
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        row = 5
        total = Decimal('0')
        
        for conta in contas:
            ws.cell(row=row, column=1, value=ExportacaoService.formatar_data(conta.get('data_vencimento')))
            ws.cell(row=row, column=2, value=conta.get('cliente_fantasia') or conta.get('cliente_nome'))
            ws.cell(row=row, column=3, value=conta.get('descricao'))
            ws.cell(row=row, column=4, value=conta.get('numero_documento') or '')
            
            valor = conta.get('valor_total', 0)
            ws.cell(row=row, column=5, value=float(valor) if valor else 0)
            ws.cell(row=row, column=5).number_format = 'R$ #,##0.00'
            
            ws.cell(row=row, column=6, value=conta.get('status', '').upper())
            ws.cell(row=row, column=7, value=conta.get('filial_nome') or '')
            ws.cell(row=row, column=8, value=conta.get('centro_custo_descricao') or '')
            
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = border
                ws.cell(row=row, column=col).alignment = Alignment(vertical='center')
            
            if valor:
                total += Decimal(str(valor))
            row += 1
        
        ws.cell(row=row, column=4, value='TOTAL:')
        ws.cell(row=row, column=4).font = total_font
        ws.cell(row=row, column=4).fill = total_fill
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='right')
        
        ws.cell(row=row, column=5, value=float(total))
        ws.cell(row=row, column=5).number_format = 'R$ #,##0.00'
        ws.cell(row=row, column=5).font = total_font
        ws.cell(row=row, column=5).fill = total_fill
        
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = border
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 20
        
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def exportar_clientes_pdf(clientes: List[Dict]) -> BytesIO:
        """Exporta clientes para PDF"""
        buffer = BytesIO()
        
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                               rightMargin=1*cm, leftMargin=1*cm,
                               topMargin=1.5*cm, bottomMargin=1*cm)
        
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#6B46C1'),
            spaceAfter=20,
            alignment=TA_CENTER
        )
        
        elements.append(Paragraph("Relatório de Clientes", title_style))
        elements.append(Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", 
                                 styles['Normal']))
        elements.append(Spacer(1, 0.5*cm))
        
        # Tabela de dados
        data = [['Nome/Razão Social', 'CNPJ/CPF', 'Telefone', 'Email', 'Cidade']]
        
        for cliente in clientes:
            data.append([
                (cliente.get('nome') or cliente.get('razao_social', ''))[:40],
                cliente.get('cnpj', '')[:18],
                cliente.get('telefone', '')[:15],
                cliente.get('email', '')[:30],
                cliente.get('cidade', '')[:20]
            ])
        
        table = Table(data, colWidths=[7*cm, 4*cm, 3.5*cm, 6*cm, 4*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6B46C1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        
        elements.append(table)
        doc.build(elements)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def exportar_clientes_excel(clientes: List[Dict]) -> BytesIO:
        """Exporta clientes para Excel"""
        buffer = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Clientes"
        
        header_fill = PatternFill(start_color="6B46C1", end_color="6B46C1", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        ws['A1'] = "Relatório de Clientes"
        ws['A1'].font = Font(bold=True, size=14, color="6B46C1")
        ws.merge_cells('A1:H1')
        
        ws['A2'] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws.merge_cells('A2:H2')
        
        headers = ['Código', 'Nome/Razão Social', 'CNPJ/CPF', 'Telefone', 'Email', 'Cidade', 'Estado', 'Tipo Serviço']
        row = 4
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        row = 5
        for cliente in clientes:
            ws.cell(row=row, column=1, value=cliente.get('codigo'))
            ws.cell(row=row, column=2, value=cliente.get('nome') or cliente.get('razao_social'))
            ws.cell(row=row, column=3, value=cliente.get('cnpj'))
            ws.cell(row=row, column=4, value=cliente.get('telefone'))
            ws.cell(row=row, column=5, value=cliente.get('email'))
            ws.cell(row=row, column=6, value=cliente.get('cidade'))
            ws.cell(row=row, column=7, value=cliente.get('estado'))
            ws.cell(row=row, column=8, value=cliente.get('tipo_servico_nome') or '')
            
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = border
                ws.cell(row=row, column=col).alignment = Alignment(vertical='center')
            row += 1
        
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 8
        ws.column_dimensions['H'].width = 25
        
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def exportar_fornecedores_pdf(fornecedores: List[Dict]) -> BytesIO:
        """Exporta fornecedores para PDF"""
        buffer = BytesIO()
        
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                               rightMargin=1*cm, leftMargin=1*cm,
                               topMargin=1.5*cm, bottomMargin=1*cm)
        
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#6B46C1'),
            spaceAfter=20,
            alignment=TA_CENTER
        )
        
        elements.append(Paragraph("Relatório de Fornecedores", title_style))
        elements.append(Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", 
                                 styles['Normal']))
        elements.append(Spacer(1, 0.5*cm))
        
        # Tabela de dados
        data = [['Nome/Razão Social', 'CNPJ/CPF', 'Telefone', 'Email', 'Cidade']]
        
        for fornecedor in fornecedores:
            data.append([
                (fornecedor.get('nome') or fornecedor.get('razao_social', ''))[:40],
                fornecedor.get('cnpj', '')[:18],
                fornecedor.get('telefone', '')[:15],
                fornecedor.get('email', '')[:30],
                fornecedor.get('cidade', '')[:20]
            ])
        
        table = Table(data, colWidths=[7*cm, 4*cm, 3.5*cm, 6*cm, 4*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6B46C1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        
        elements.append(table)
        doc.build(elements)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def exportar_fornecedores_excel(fornecedores: List[Dict]) -> BytesIO:
        """Exporta fornecedores para Excel"""
        buffer = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Fornecedores"
        
        header_fill = PatternFill(start_color="6B46C1", end_color="6B46C1", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        ws['A1'] = "Relatório de Fornecedores"
        ws['A1'].font = Font(bold=True, size=14, color="6B46C1")
        ws.merge_cells('A1:H1')
        
        ws['A2'] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws.merge_cells('A2:H2')
        
        headers = ['Código', 'Nome/Razão Social', 'CNPJ/CPF', 'Telefone', 'Email', 'Cidade', 'Estado', 'Tipo Serviço']
        row = 4
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        row = 5
        for fornecedor in fornecedores:
            ws.cell(row=row, column=1, value=fornecedor.get('codigo'))
            ws.cell(row=row, column=2, value=fornecedor.get('nome') or fornecedor.get('razao_social'))
            ws.cell(row=row, column=3, value=fornecedor.get('cnpj'))
            ws.cell(row=row, column=4, value=fornecedor.get('telefone'))
            ws.cell(row=row, column=5, value=fornecedor.get('email'))
            ws.cell(row=row, column=6, value=fornecedor.get('cidade'))
            ws.cell(row=row, column=7, value=fornecedor.get('estado'))
            ws.cell(row=row, column=8, value=fornecedor.get('tipo_servico_nome') or '')
            
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = border
                ws.cell(row=row, column=col).alignment = Alignment(vertical='center')
            row += 1
        
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 8
        ws.column_dimensions['H'].width = 25
        
        wb.save(buffer)
        buffer.seek(0)
        return buffer

